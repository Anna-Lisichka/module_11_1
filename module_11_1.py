
import openpyxl as op
import pandas as pd
from openpyxl.chart import BarChart, Reference


'''Мой выбор пал на библиотеку, способную работать с файлами excel. Она поименована в задании,
является для меня новой и, возможно, будет полезной в дальнейшем, так почему бы не начать 
ее изучение сейчас, совместив необходимость с полезностью...
Для корректного чтения файлов excel было необходимо установить дополнительные библиотеки'''

''' 1. Библиотека "pandas". Для считывания файла excel используется функция pandas.read_excel()
Данная функция считывает файл excel в DataFrame (двухмерная структура данных), вывод в виде таблицы

2. "xlrd" - это внутренняя библиотека используется для чтения и форматирования книг Excel. 
Чтение файлов excel без установки данной библиотеки приводит к ошибке.
Однако xlrd поддерживает только стандартный формат файлов - .xls.

3. "OpenPyXL" - это библиотека для чтения/записи файлов Excel 2010 и более поздних версий
с расширениями xlsx/xlsm/xltx/xltm'''



# считываем файл excel. Функция read_excel() библиотеки pandas. По умолчанию - это лист1
# чтобы считать другой лист, его надо указать по номеру, начиная с 0 (sheet_name=1)
# или названию (sheet_name='Лист1'), для считывания всех листов книги - None
balance_sheet = pd.read_excel('./files/balance.xls', sheet_name=None)
print(balance_sheet)


# класс pandas.ExcelFile() используется для облегчения работы с несколькими листами одного файла.
# файл будет читаться в память только один раз.
file_location = './files/plan.xls'
p_sheet = pd.ExcelFile(file_location)

# Печатаем название листов в данном файле
print(p_sheet.sheet_names)

# сохраняем таблицу указанного листа книги excel в переменную и выводим в консоль
plan_materials = pd.read_excel(p_sheet, 'Sheet1')
print(plan_materials)


# можно использовать в качестве контекстного менеджера
# Но основным вариантом использования данного класса является синтаксический анализ нескольких листов
# с разными параметрами. Когда данные листов отличаются
data = {}
with pd.ExcelFile('./files/plan_2.xls') as file:
    data['Sheet1'] = pd.read_excel(file, "Sheet1", index_col=None, na_values=["NA"])
    # data["Sheet2"] = pd.read_excel(file, "Sheet2", index_col=None, na_values=["NA"])
    data["Sheet2"] = pd.read_excel(file, "Sheet2", index_col=1)
print(data)

# эквивалент с использованием списка имен листов и указанием колонки, которая будет использоваться
# в качестве индекса
data_2 = pd.read_excel(file, ["Sheet1", "Sheet2"], index_col=0, na_values=["NA"])
print(data_2)

# Возвращает 1-й и 2-й листы в виде словаря `DataFrame`.
print(pd.read_excel('./files/plan_2.xls', sheet_name=["Sheet1", 1]))



# Функция load_workbook() библиотеки OpenPyXL принимает имя файла в качестве аргумента
# и возвращает объект рабочей книги, который представляет файл
wb = op.load_workbook('./files/graph.xlsx')

# выводим имя книги excel
print(wb.sheetnames)

# извлекаем значения из определенных ячеек с определенного листа книги excel
sheet = wb.sheetnames[0]
print(sheet)


# Построение диаграммы с использованием библиотеки OpenPyXL

wb = op.Workbook()
ws = wb.active

# Создание отдельного файла
data = [["Product", "Online", "Store"],
    [1, 30, 45],
    [2, 40, 30],
    [3, 40, 25],
    [4, 50, 30],
    [5, 30, 25],
    [6, 25, 35],
    [7, 20, 40],]
for row in data:
    ws.append(row)

# выбираем диапазоны значений для диаграммы
values = Reference(worksheet=ws,
                 min_row=1,
                 max_row=8,
                 min_col=2,
                 max_col=3)
# создаем объект столбчатой диаграммы
chart = BarChart()
# добавляем в диаграмму выбранный диапазон значений
chart.add_data(values, titles_from_data=True)
# привязываем диаграмму к ячейке `E15`
ws.add_chart(chart, "A11")
# определяем размеры диаграммы в сантиметрах
chart.width = 20
chart.height = 5
# сохраняем и смотрим что получилось
wb.save("./files/graph_1.xlsx")

wb = op.load_workbook('./files/graph_1.xlsx')
# список рабочих листов
name_sheet = wb.sheetnames
print(name_sheet)

# получим доступ к рабочему листу 'Sheet1' используя его индекс в списке `wb.sheetnames`
sheet = wb[name_sheet[0]]
print(sheet)






